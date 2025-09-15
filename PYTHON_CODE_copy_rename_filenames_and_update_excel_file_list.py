#Remove film number and change /add name to excel sheet
#   -copies files and then edits the copies
#   -removes film numbers, and removes spaces replacing them with '_' 
#   -adds sequencing (_seq1, _seq2, etc.)
#   -edits excel doc (list of original file names) with the modified name in the modified column


import os
import glob #to finds files and directories matching specified patterns
import shutil  #for moving files to new locations
import pandas as pd  #for excel doc use
import re #for renaming, removing '_' and ' ' and segmenting into parts (e.g. filenames like PP01_k 19930231 )

from pathlib import Path
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

print('')
 
# specify the source and target directory, target directory for file copies to manipulate, leaving originals unchanged for redundancy
source_dir = Path(r"")  #ADD PATH in ""
target_dir = Path(r"")  #ADD PATH in ""

#excel path and file to list of filenames with their filepaths
excel_path = Path(r"") #ADD PATH in ""
wb = load_workbook(excel_path) #load file/make active to edit
ws = wb.active

# loop through excel rows matching filenames in source directory, note: column 1 has original filenames, column 2 has modified filenames
for row in ws.iter_rows(min_row=2): #'2' to skip header row
    original_cell = row[0]
    modified_cell = row[1]
    original_name = original_cell.value
    
    if not original_name: #skip if no match to next row
        continue

    original_name = original_name.strip() #note: .strip() deletes leading or trailing whitespaces incase that is an issue
    source_path = os.path.join(source_dir, original_name)
    
    #copy files into target directory by changing the path to edit leaving originals alone
    filename = os.path.basename(source_path)
    copied_path = os.path.join(target_dir, filename)
    shutil.copy2(source_path, copied_path)  # copy with metadata (timestamps etc.)
         
    new_path = copied_path  

    #remove film numbers by splitting up filenames by spaces and identifying ones that start with F0, K0, K1 and removing those segments/splits
    if splitted and splitted[0].startswith(("F0", "K0", "K1")): #for files that have a film number currently
        base_parts = splitted[1:] #remove film number, the [1:] slice skips the first element.  if splitted = ["F02322", "Fog", "1973"], then base_parts = ["Fog", "1973"].
        base_name = "_".join(base_parts) #join base_parts list into a single string, with underscores _ between the parts and not spaces.
        new_name = f"{base_name}{ext}" #adds file extension back to filename
        new_path = os.path.join(target_dir, new_name) #This builds the full path to where the renamed file will go

        counter = 1 #sequencing
        while os.path.exists(new_path):  # Using os.path.exists for strings. Add in sequencing to ALL files, add counter to sequence and extension at end to complete filename
            new_name = f"{base_name}_seq{counter}{ext}"
            new_path = target_dir / new_name
            counter += 1
        
        print("Renaming: ", os.path.basename(copied_path), " -> ", new_name) #print updated names if changed
        os.rename(copied_path, new_path) 
    else:
        print("Skipped renaming: ", os.path.basename(copied_path)) #print name of files not renamed

    # update excel file with final filename 
    modified_cell.value = os.path.basename(new_path)

# Save excel document 
wb.save(excel_path)

print("Done. Files copied, film numbers and spaces removed, underscores and sequencing added. Excel file updated.")

